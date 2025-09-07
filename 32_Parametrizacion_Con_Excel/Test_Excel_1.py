import re
import time
from playwright.sync_api import Playwright, sync_playwright, expect
from AA_Funciones_Globales_Total.Funciones_Globales_Total import Funciones_Globales_Total #para llamar las funciones viene de la carpeta y funcion
import random
import pytest
import sys #Libreria para informacion directamente del sistema operativo
import openpyxl

tiempo=0.3
ruta="Imagenes_Globales/Select"
pdf1="D:/Practica_Playwright/18_Unpload_Files/Imagenes/Upload/Playwright_Prueba_Upload.jpg"
rutaexcel="D:/Practica_Playwright/32_Parametrizacion_Con_Excel/Playwright.xlsx"
#invertir la diagonal, \ por /
registros=8

#Leyendo el archivo de Excel
archivo=openpyxl.load_workbook(rutaexcel)
def Numero_Filas(hoja):
    ac=archivo[hoja]
    return ac.max_row

def Dato_Columna(hoja,fila,col):
    ac=archivo[hoja]
    col=ac.cell(int(fila),int(col))
    return col.value

print(Numero_Filas("Hoja1"))
print(Dato_Columna("Hoja1",2,1))

