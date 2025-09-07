import re
import time
from playwright.sync_api import Playwright, sync_playwright, expect
from AA_Funciones_Globales_Total.Funciones_Globales_Total import Funciones_Globales_Total #para llamar las funciones viene de la carpeta y funcion
import random
import pytest
import sys #Libreria para informacion directamente del sistema operativo
import openpyxl

tiempo=0.3
ruta="Imagenes_Globales/Select"
pdf1="D:/Practica_Playwright/18_Unpload_Files/Imagenes/Upload/Playwright_Prueba_Upload.jpg"
rutaexcel="D:/Practica_Playwright/32_Parametrizacion_Con_Excel/Playwright.xlsx"
#invertir la diagonal, \ por /
registros=8

#Leyendo el archivo de Excel
archivo=openpyxl.load_workbook(rutaexcel)
def Numero_Filas(hoja):
    ac=archivo[hoja]
    return ac.max_row

def Dato_Columna(hoja,fila,col):
    ac=archivo[hoja]
    col=ac.cell(int(fila),int(col))
    return col.value

print(Numero_Filas("Hoja1"))
print(Dato_Columna("Hoja1",2,1))

def test_excel1(set_up_Excel) ->None:
    page=set_up_Excel
    F=Funciones_Globales_Total(page)
    F.Validar_Titulo_PaginaWeb("Formulario de Ejemplo")
    F.Scroll_x_y(0,400)
    print("Cargando Excel")
    F.Esperar(2)
    
    
    for n in range(2,registros):
        Nombre=Dato_Columna("Hoja1",n,1)
        print(Nombre)
        Apellido=Dato_Columna("Hoja1",n,2)
        print(Apellido)
        Email=Dato_Columna("Hoja1",n,3)
        print(Email)
        Tlf=Dato_Columna("Hoja1",n,4)
        print(Tlf)
        Direccion=Dato_Columna("Hoja1",n,5)
        print(Direccion)
     
    
    F.Texto("//input[@id='nombre']",Nombre)
    F.Texto("//input[@id='apellidos']",Apellido)
    F.Texto("//input[@id='tel']",str(Tlf))
    F.Texto("//input[@id='email']",Email)
    F.Texto("//input[@id='direccion']",Direccion)
    F.Esperar(1)
    page.goto("https://validaciones.rodrigovillanueva.com.mx/Form1.html")
    #Lo regresamos a la pag, para que pueda seguir cargando la informacion
    F.Esperar(1)
    page.goto("https://validaciones.rodrigovillanueva.com.mx/Form1.html")
    #Lo regresamos a la pag, para que pueda seguir cargando la informacion
    F.Esperar(1)
    